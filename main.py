from time import sleep
import re
from openpyxl import Workbook

from selenium.webdriver import Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


navegador = Chrome()
navegador.get("https://justicaaberta.cnj.jus.br/")
navegador.maximize_window()

wait = WebDriverWait(navegador, 20)

emails = set()


def click_js(elemento):
    navegador.execute_script("arguments[0].click();", elemento)


def esperar_opcoes():
    wait.until(
        lambda d: len(
            d.find_elements(By.CSS_SELECTOR, "div.option")
        ) > 5
    )


def coletar_tabela():
    print("Coletando tabela...")

    wait.until(
        lambda d: len(
            d.find_elements(By.CSS_SELECTOR, "table tbody tr")
        ) > 0
    )

    paginas_visitadas = set()

    while True:
        linhas = navegador.find_elements(By.CSS_SELECTOR, "table tbody tr")

        texto_pagina = "\n".join([linha.text for linha in linhas])

        if texto_pagina in paginas_visitadas:
            print("Página repetida. Encerrando paginação.")
            break

        paginas_visitadas.add(texto_pagina)

        for linha in linhas:
            texto_linha = linha.text

            # coleta somente cartórios ativos
            if "(Ativo)" not in texto_linha:
                continue

            encontrados = re.findall(
                r"[\w\.-]+@[\w\.-]+\.\w+",
                texto_linha
            )

            emails.update(encontrados)

        print("Emails encontrados:", len(emails))

        botoes = navegador.find_elements(By.TAG_NAME, "button")
        proximo = None

        for botao in botoes:
            if not botao.is_displayed():
                continue

            texto = botao.text.strip()
            html = botao.get_attribute("innerHTML") or ""

            if texto == ">" or "&gt;" in html or ">" in texto:
                proximo = botao
                break

        if proximo is None:
            print("Não encontrou botão próxima")
            break

        classe = proximo.get_attribute("class") or ""
        disabled = proximo.get_attribute("disabled")
        aria_disabled = proximo.get_attribute("aria-disabled")

        if (
            "disabled" in classe.lower()
            or disabled is not None
            or aria_disabled == "true"
        ):
            print("Fim da paginação")
            break

        click_js(proximo)
        sleep(2)


def salvar_excel():
    wb = Workbook()
    ws = wb.active

    ws.title = "Emails"
    ws["A1"] = "Email"
    ws.column_dimensions["A"].width = 50

    for linha, email in enumerate(sorted(emails), start=2):
        ws.cell(row=linha, column=1, value=email)

    arquivo = "emails_cartorios_ativos.xlsx"
    wb.save(arquivo)

    print(f"\nArquivo gerado: {arquivo}")
    print(f"Total de emails exportados: {len(emails)}")


# =====================================================
# ESTADOS
# =====================================================

click_estado = wait.until(
    EC.element_to_be_clickable(
        (
            By.XPATH,
            '//*[@id="app"]/main/div[1]/div[4]/div/div/div[2]/div[1]/div[2]/div[2]'
        )
    )
)

click_js(click_estado)
sleep(2)
esperar_opcoes()

estados = navegador.find_elements(By.CSS_SELECTOR, "div.option")
total_estados = len(estados)


for indice_estado in range(total_estados):

    click_estado = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                '//*[@id="app"]/main/div[1]/div[4]/div/div/div[2]/div[1]/div[2]/div[2]'
            )
        )
    )

    click_js(click_estado)
    sleep(1)
    esperar_opcoes()

    estados = navegador.find_elements(By.CSS_SELECTOR, "div.option")

    if indice_estado >= len(estados):
        break

    estado = estados[indice_estado]

    nome_estado = estado.get_attribute("innerText").strip()

    if len(nome_estado) != 2:
        continue

    print("\n===================")
    print("Estado:", nome_estado)
    print("===================")

    click_js(estado)
    sleep(1)

    botao_pesquisar = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                '//*[@id="app"]/main/div[1]/div[4]/div/div/div[2]/div[2]/button'
            )
        )
    )

    click_js(botao_pesquisar)
    sleep(3)

    coletar_tabela()

    print("Total parcial:", len(emails))


print("\n================")
print("FINAL")
print("================")

for email in sorted(emails):
    print(email)

print("\nTotal:", len(emails))

salvar_excel()

navegador.quit()